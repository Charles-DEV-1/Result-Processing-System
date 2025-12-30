"""
Maranatha University Result Processing System
Flask Application with SQLAlchemy
"""

from flask import Flask, render_template_string, request, redirect, url_for, flash, session, send_file, render_template
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///results.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ============= DATABASE MODELS =============

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'teacher' or 'admin'
    full_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    matric_number = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    department = db.Column(db.String(100), nullable=False)
    level = db.Column(db.String(10), nullable=False)
    results = db.relationship('Result', backref='student', lazy=True)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_code = db.Column(db.String(20), unique=True, nullable=False)
    course_title = db.Column(db.String(200), nullable=False)
    course_unit = db.Column(db.Integer, nullable=False)
    results = db.relationship('Result', backref='course', lazy=True)

class Result(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    semester = db.Column(db.String(20), nullable=False)
    session = db.Column(db.String(20), nullable=False)
    ca_score = db.Column(db.Float, nullable=False)
    exam_score = db.Column(db.Float, nullable=False)
    total_score = db.Column(db.Float, nullable=False)
    grade = db.Column(db.String(2), nullable=False)
    grade_point = db.Column(db.Float, nullable=False)
    entered_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ============= HELPER FUNCTIONS =============

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def calculate_grade(total_score):
    if total_score >= 70:
        return 'A', 5.0
    elif total_score >= 60:
        return 'B', 4.0
    elif total_score >= 50:
        return 'C', 3.0
    elif total_score >= 45:
        return 'D', 2.0
    elif total_score >= 40:
        return 'E', 1.0
    else:
        return 'F', 0.0

def calculate_gpa(results):
    if not results:
        return 0.0
    total_points = sum(r.grade_point * r.course.course_unit for r in results)
    total_units = sum(r.course.course_unit for r in results)
    return round(total_points / total_units, 2) if total_units > 0 else 0.0


# ============= ROUTES =============

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))  
    return render_template("index.html")

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        role = request.form.get('role', 'teacher')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('register'))
        
        hashed_password = generate_password_hash(password)
        user = User(username=username, password=hashed_password, full_name=full_name, role=role)
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login', 'success')
        return redirect(url_for('login'))
    return render_template("register.html")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template("login.html")

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    total_students = Student.query.count()
    total_courses = Course.query.count()
    total_results = Result.query.count()
    return render_template('dashboard.html', total_students=total_students, 
                           total_courses=total_courses, total_results=total_results)

@app.route('/enter_result', methods=['GET', 'POST'])
@login_required
def enter_result():
    if request.method == 'POST':
        # Get or create student
        matric = request.form.get('matric_number')
        student = Student.query.filter_by(matric_number=matric).first()
        
        if not student:
            student = Student(
                matric_number=matric,
                full_name=request.form.get('student_name'),
                department=request.form.get('department'),
                level=request.form.get('level')
            )
            db.session.add(student)
            db.session.flush()
        
        # Get or create course
        course_code = request.form.get('course_code')
        course = Course.query.filter_by(course_code=course_code).first()
        
        if not course:
            course = Course(
                course_code=course_code,
                course_title=request.form.get('course_title'),
                course_unit=int(request.form.get('course_unit'))
            )
            db.session.add(course)
            db.session.flush()
        
        # Calculate scores
        ca_score = float(request.form.get('ca_score'))
        exam_score = float(request.form.get('exam_score'))
        total_score = ca_score + exam_score
        grade, grade_point = calculate_grade(total_score)
        
        # Check for duplicate
        existing = Result.query.filter_by(
            student_id=student.id,
            course_id=course.id,
            semester=request.form.get('semester'),
            session=request.form.get('session')
        ).first()
        
        if existing:
            flash('Result already exists for this student, course, and semester', 'error')
            return redirect(url_for('enter_result'))
        
        # Save result
        result = Result(
            student_id=student.id,
            course_id=course.id,
            semester=request.form.get('semester'),
            session=request.form.get('session'),
            ca_score=ca_score,
            exam_score=exam_score,
            total_score=total_score,
            grade=grade,
            grade_point=grade_point,
            entered_by=session['user_id']
        )
        db.session.add(result)
        db.session.commit()
        
        flash('Result entered successfully!', 'success')
        return redirect(url_for('enter_result'))
    
    courses = Course.query.all()
    return render_template('enter_result.html', courses=courses)

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():

    results = (
        db.session.query(Result)
        .join(Student)
        .join(Course)
        .all()
    )

    students_map = {}

    for r in results:
        sid = r.student.id

        if sid not in students_map:
            students_map[sid] = {
                "student": r.student,
                "courses": {},
                "gpa": 0
            }

        # one course → one column (TOTAL only)
        students_map[sid]["courses"][r.course.course_code] = r.total_score

    # calculate GPA per student
    for data in students_map.values():
        student_results = Result.query.filter_by(
            student_id=data["student"].id
        ).all()
        data["gpa"] = calculate_gpa(student_results)

    return render_template(
        "admin_dashboard.html",
        students=list(students_map.values())
    )


@app.route('/admin/courses', methods=['GET', 'POST'])
@admin_required
def manage_courses():
    if request.method == 'POST':
        course = Course(
            course_code=request.form.get('course_code'),
            course_title=request.form.get('course_title'),
            course_unit=int(request.form.get('course_unit'))
        )
        db.session.add(course)
        db.session.commit()
        flash('Course added successfully!', 'success')
        return redirect(url_for('manage_courses'))
    
    courses = Course.query.all()
    
    return render_template('manage_courses.html', courses=courses)

@app.route('/export/excel')
@admin_required
def export_excel():
    level = request.args.get('level', '')
    department = request.args.get('department', '')
    semester = request.args.get('semester', '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"

    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # STEP 1: Get filtered results
    query = (
        db.session.query(Result)
        .join(Student)
        .join(Course)
    )

    if level:
        query = query.filter(Student.level == level)
    if department:
        query = query.filter(Student.department == department)
    if semester:
        query = query.filter(Result.semester == semester)

    results = query.all()

    # STEP 2: Group results by student
    students_map = {}
    all_courses = set()

    for r in results:
        sid = r.student.id
        course_code = r.course.course_code
        all_courses.add(course_code)

        if sid not in students_map:
            students_map[sid] = {
                "student": r.student,
                "courses": {}
            }

        students_map[sid]["courses"][course_code] = r.total_score

    # STEP 3: Headers
    headers = ["Matric No", "Student Name", "Department", "Level"]
    headers.extend(sorted(all_courses))
    headers.append("GPA")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # STEP 4: Fill rows
    row = 2
    for data in students_map.values():
        student = data["student"]

        student_results = Result.query.filter_by(student_id=student.id)
        if semester:
            student_results = student_results.filter_by(semester=semester)

        student_results = student_results.all()
        gpa = calculate_gpa(student_results)

        ws.cell(row=row, column=1, value=student.matric_number)
        ws.cell(row=row, column=2, value=student.full_name)
        ws.cell(row=row, column=3, value=student.department)
        ws.cell(row=row, column=4, value=student.level)

        col_index = 5
        for course in sorted(all_courses):
            ws.cell(
                row=row,
                column=col_index,
                value=data["courses"].get(course, "-")
            )
            col_index += 1

        ws.cell(row=row, column=col_index, value=gpa)
        row += 1

    # STEP 5: Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Results_{department}_{level}_{semester}.xlsx".replace(" ", "_")

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# ============= INITIALIZE DATABASE =============

def init_db():
    with app.app_context():
        db.create_all()
        
        # Create default admin if not exists
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password=generate_password_hash('admin123'),
                full_name='System Administrator',
                role='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print("Default admin created - username: admin, password: admin123")

if __name__ == '__main__':
    init_db()
    app.run(debug=False)